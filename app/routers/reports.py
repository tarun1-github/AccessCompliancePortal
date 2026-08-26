import csv, io
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.models.engineer import Engineer
from app.models.applications import Application
from app.models.verification import EngineerApplicationAccess

router = APIRouter(prefix="/api/reports", tags=["reports"])

def rows(db):
    data = db.query(EngineerApplicationAccess, Engineer, Application).join(Engineer, Engineer.id==EngineerApplicationAccess.engineer_id).join(Application, Application.id==EngineerApplicationAccess.application_id).order_by(Engineer.name, Application.name).all()
    for access, engineer, app in data:
        yield [engineer.name, engineer.alias, engineer.email or "", engineer.rm_email or "", engineer.role, app.name, access.access_status, access.verification_status, access.arm_ticket or "", access.ticket_status or "Not Started", access.remarks or "", access.last_verified_date.isoformat() if access.last_verified_date else ""]

HEADERS=["User","Alias","User Email","Supervisor Email","Role","Application","Access Status","Verification Status","ARM Ticket","Ticket Status","Remarks","Last Verified"]

@router.get('/csv')
def export_csv(db: Session = Depends(get_db)):
    stream=io.StringIO(); w=csv.writer(stream); w.writerow(HEADERS); [w.writerow(r) for r in rows(db)]
    return StreamingResponse(iter([stream.getvalue()]), media_type='text/csv', headers={'Content-Disposition':'attachment; filename=access_verification_report.csv'})

@router.get('/excel')
def export_excel(db: Session = Depends(get_db)):
    wb=Workbook(); ws=wb.active; ws.title='Access Verification'; ws.append(HEADERS)
    for cell in ws[1]:
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='2563EB'); cell.alignment=Alignment(horizontal='center')
    for row in rows(db): ws.append(row)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns:
        letter=col[0].column_letter; ws.column_dimensions[letter].width=min(max(max(len(str(c.value or '')) for c in col)+2,12),42)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=access_verification_report.xlsx'})
