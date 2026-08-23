from fastapi import APIRouter, Depends, HTTPException, Request

from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from sqlalchemy.orm import Session

import csv
import io

from app.database import get_db
from app.models.engineer import Engineer
from app.models.applications import Application
from app.models.verification import EngineerApplicationAccess
from app.models.settings import PortalSetting


router = APIRouter()

templates = Jinja2Templates(
    directory="app/templates"
)


# ============================================================
# CONSTANTS
# ============================================================

DEFAULT_REMINDER_DAYS = 15

ARM_REQUEST_NOT_INITIATED = "Request Not Initiated"
ARM_NOT_REQUIRED = "Not Required"
ARM_APPROVAL_PENDING = "Approval Pending"
ARM_COMPLETED = "Completed"

VALID_ARM_STATUSES = {
    ARM_REQUEST_NOT_INITIATED,
    ARM_NOT_REQUIRED,
    ARM_APPROVAL_PENDING,
    ARM_COMPLETED,
}


# ============================================================
# NORMALIZE VERIFICATION STATUS
# ============================================================

def normalize_verification_status(value):

    if not value:
        return "Pending"

    value = str(value).strip()

    if value in {
        "Verified",
        "Issue",
        "Pending",
    }:
        return value

    return "Pending"


# ============================================================
# NORMALIZE ARM STATUS
#
# IMPORTANT:
# Default ARM status is NOT REQUIRED.
# Old "Pending" values are treated as Approval Pending.
# ============================================================

def normalize_ticket_status(value):

    if not value:
        return ARM_NOT_REQUIRED

    value = str(value).strip()

    if not value:
        return ARM_NOT_REQUIRED

    if value.lower() == "pending":
        return ARM_APPROVAL_PENDING

    if value in VALID_ARM_STATUSES:
        return value

    return ARM_NOT_REQUIRED


# ============================================================
# BUILD ENGINEER APPLICATION DATA
# ============================================================

def build_engineer_applications(
    engineer_id: int,
    db: Session,
):

    rows = (
        db.query(
            EngineerApplicationAccess,
            Application,
        )
        .join(
            Application,
            Application.id
            == EngineerApplicationAccess.application_id,
        )
        .filter(
            EngineerApplicationAccess.engineer_id
            == engineer_id,

            Application.active == True,
        )
        .order_by(
            Application.name
        )
        .all()
    )

    applications = []

    for access, application in rows:

        applications.append(
            {
                "access_id": access.id,

                "application_id": application.id,

                "application_name": application.name,

                "access_status":
                    access.access_status,

                "verification_status":
                    normalize_verification_status(
                        access.verification_status
                    ),

                "arm_ticket":
                    access.arm_ticket or "",

                "ticket_status":
                    normalize_ticket_status(
                        access.ticket_status
                    ),

                "last_verified_date":
                    (
                        access.last_verified_date.isoformat()
                        if access.last_verified_date
                        else None
                    ),

                "remarks":
                    access.remarks or "",
            }
        )

    return applications


# ============================================================
# REMINDER DAYS
#
# Compatible with multiple PortalSetting schemas.
# ============================================================

def reminder_days(db: Session):

    try:

        columns = PortalSetting.__table__.columns.keys()

        setting = None

        if "key" in columns:

            setting = (
                db.query(PortalSetting)
                .filter(
                    PortalSetting.key
                    == "reminder_days"
                )
                .first()
            )

        elif "setting_key" in columns:

            setting = (
                db.query(PortalSetting)
                .filter(
                    PortalSetting.setting_key
                    == "reminder_days"
                )
                .first()
            )

        if setting is not None:

            value = None

            if hasattr(setting, "value"):
                value = setting.value

            elif hasattr(setting, "setting_value"):
                value = setting.setting_value

            if value is not None:

                value = str(value).strip()

                if value.isdigit():

                    days = int(value)

                    if days > 0:
                        return days

    except Exception:

        db.rollback()

    return DEFAULT_REMINDER_DAYS


# ============================================================
# GET ACTIVE ENGINEER
# ============================================================

def get_active_engineer_or_404(
    engineer_id: int,
    db: Session,
):

    engineer = (
        db.query(Engineer)
        .filter(
            Engineer.id == engineer_id,
            Engineer.active == True,
        )
        .first()
    )

    if engineer is None:

        raise HTTPException(
            status_code=404,
            detail="Engineer not found",
        )

    return engineer


# ============================================================
# GET SUPERVISOR
# ============================================================

def get_supervisor_or_403(
    supervisor_id: int,
    db: Session,
):

    supervisor = (
        db.query(Engineer)
        .filter(
            Engineer.id == supervisor_id,
            Engineer.active == True,
        )
        .first()
    )

    if supervisor is None:

        raise HTTPException(
            status_code=403,
            detail="Valid supervisor not found",
        )

    if (
        str(
            supervisor.role or ""
        ).upper()
        != "SUPERVISOR"
    ):

        raise HTTPException(
            status_code=403,
            detail="Valid supervisor not found",
        )

    return supervisor


# ============================================================
# CHECK ENGINEER BELONGS TO SUPERVISOR
#
# Supports both:
# supervisor_id
# and RM email mapping.
# ============================================================

def engineer_belongs_to_supervisor(
    engineer: Engineer,
    supervisor: Engineer,
):

    columns = Engineer.__table__.columns.keys()

    if "supervisor_id" in columns:

        supervisor_id = getattr(
            engineer,
            "supervisor_id",
            None,
        )

        if supervisor_id is not None:

            return (
                int(supervisor_id)
                == int(supervisor.id)
            )

    if "rm_email" in columns:

        engineer_rm_email = (
            getattr(
                engineer,
                "rm_email",
                None,
            )
            or ""
        ).strip().lower()

        supervisor_email = (
            getattr(
                supervisor,
                "email",
                None,
            )
            or ""
        ).strip().lower()

        if (
            engineer_rm_email
            and supervisor_email
        ):

            return (
                engineer_rm_email
                == supervisor_email
            )

    # Existing installations may not yet have
    # a supervisor mapping column.
    # Allow the operation rather than blocking ARM save.
    return True


# ============================================================
# PAGE FOR ENGINEER / SUPERVISOR
# ============================================================

def page_for(
    engineer: Engineer,
    request: Request,
    db: Session,
):

    if (
        str(
            engineer.role or ""
        ).upper()
        == "SUPERVISOR"
    ):

        engineers = (
            db.query(Engineer)
            .filter(
                Engineer.active == True,
                Engineer.id != engineer.id,
            )
            .order_by(
                Engineer.name
            )
            .all()
        )

        return templates.TemplateResponse(
            request=request,
            name="supervisor.html",
            context={
                "request": request,
                "engineer": engineer,
                "engineers": engineers,
                "reminder_days": reminder_days(db),
            },
        )

    return templates.TemplateResponse(
        request=request,
        name="verification.html",
        context={
            "request": request,
            "engineer": engineer,
            "engineer_id": engineer.id,
            "applications":
                build_engineer_applications(
                    engineer.id,
                    db,
                ),
            "reminder_days":
                reminder_days(db),
        },
    )


# ============================================================
# HOME
# ============================================================

@router.get(
    "/",
    response_class=HTMLResponse,
)
def home():

    return HTMLResponse(
        """
        <h2>CMS BOA EV Access Compliance Portal</h2>
        <p>Open your personal verification link.</p>
        """
    )


# ============================================================
# VERIFY TOKEN
# ============================================================

@router.get(
    "/verify/{token}",
    response_class=HTMLResponse,
)
def verify_engineer(
    token: str,
    request: Request,
    db: Session = Depends(get_db),
):

    engineer = (
        db.query(Engineer)
        .filter(
            Engineer.verification_token == token,
            Engineer.active == True,
        )
        .first()
    )

    if engineer is None:

        raise HTTPException(
            status_code=404,
            detail="Invalid or expired verification link",
        )

    return page_for(
        engineer,
        request,
        db,
    )


# ============================================================
# NORMAL ENGINEER DETAILS
# ============================================================

@router.get(
    "/api/engineer/{engineer_id}"
)
def get_engineer_details(
    engineer_id: int,
    db: Session = Depends(get_db),
):

    engineer = get_active_engineer_or_404(
        engineer_id,
        db,
    )

    return {
        "engineer_id": engineer.id,
        "engineer_name": engineer.name,
        "level": engineer.level,
        "email": engineer.email,
        "alias": engineer.alias,
        "role": engineer.role,
        "applications":
            build_engineer_applications(
                engineer.id,
                db,
            ),
    }


# ============================================================
# SUPERVISOR ENGINEER LIST
# ============================================================

@router.get(
    "/api/supervisor/engineers"
)
def all_engineers(
    db: Session = Depends(get_db),
):

    engineers = (
        db.query(Engineer)
        .filter(
            Engineer.active == True
        )
        .order_by(
            Engineer.name
        )
        .all()
    )

    result = []

    for engineer in engineers:

        if (
            str(
                engineer.role or ""
            ).upper()
            == "SUPERVISOR"
        ):
            continue

        applications = build_engineer_applications(
            engineer.id,
            db,
        )

        result.append(
            {
                "id": engineer.id,
                "name": engineer.name,
                "alias": engineer.alias,
                "level": engineer.level,
                "email": engineer.email,
                "role": engineer.role,
                "total": len(applications),
            }
        )

    return result


# ============================================================
# SUPERVISOR SELECTED ENGINEER DETAILS
# ============================================================

@router.get(
    "/api/supervisor/engineer/{engineer_id}"
)
def supervisor_engineer_details(
    engineer_id: int,
    db: Session = Depends(get_db),
):

    engineer = get_active_engineer_or_404(
        engineer_id,
        db,
    )

    applications = build_engineer_applications(
        engineer.id,
        db,
    )

    total = len(applications)

    verified = sum(
        1
        for x in applications
        if x["verification_status"] == "Verified"
    )

    pending = sum(
        1
        for x in applications
        if x["verification_status"] == "Pending"
    )

    issues = sum(
        1
        for x in applications
        if x["verification_status"] == "Issue"
    )

    request_not_initiated = sum(
        1
        for x in applications
        if x["ticket_status"]
        == ARM_REQUEST_NOT_INITIATED
    )

    not_required = sum(
        1
        for x in applications
        if x["ticket_status"]
        == ARM_NOT_REQUIRED
    )

    approval_pending = sum(
        1
        for x in applications
        if x["ticket_status"]
        == ARM_APPROVAL_PENDING
    )

    completed = sum(
        1
        for x in applications
        if x["ticket_status"]
        == ARM_COMPLETED
    )

    return {
        "engineer_id": engineer.id,
        "engineer_name": engineer.name,
        "level": engineer.level,
        "email": engineer.email,
        "alias": engineer.alias,
        "role": engineer.role,
        "applications": applications,

        "summary": {
            "total": total,
            "verified": verified,
            "pending": pending,
            "issues": issues,

            "request_not_initiated":
                request_not_initiated,

            "not_required":
                not_required,

            "approval_pending":
                approval_pending,

            "completed":
                completed,

            "tickets_completed":
                completed,
        },
    }


# ============================================================
# DASHBOARD
# ============================================================

@router.get(
    "/api/dashboard"
)
def dashboard(
    db: Session = Depends(get_db),
):

    records = (
        db.query(
            EngineerApplicationAccess
        )
        .all()
    )

    total = len(records)

    verified = 0
    pending = 0
    issues = 0

    request_not_initiated = 0
    not_required = 0
    approval_pending = 0
    completed = 0

    for record in records:

        verification_status = (
            normalize_verification_status(
                record.verification_status
            )
        )

        ticket_status = (
            normalize_ticket_status(
                record.ticket_status
            )
        )

        if verification_status == "Verified":
            verified += 1

        elif verification_status == "Issue":
            issues += 1

        else:
            pending += 1

        if ticket_status == ARM_REQUEST_NOT_INITIATED:
            request_not_initiated += 1

        elif ticket_status == ARM_NOT_REQUIRED:
            not_required += 1

        elif ticket_status == ARM_APPROVAL_PENDING:
            approval_pending += 1

        elif ticket_status == ARM_COMPLETED:
            completed += 1

    return {
        "total": total,
        "verified": verified,
        "pending": pending,
        "issues": issues,

        "request_not_initiated":
            request_not_initiated,

        "not_required":
            not_required,

        "approval_pending":
            approval_pending,

        "completed":
            completed,

        "ticket_completed":
            completed,

        "ticket_approval_pending":
            approval_pending,

        "ticket_pending":
            approval_pending,

        "reminder_days":
            reminder_days(db),
    }


# ============================================================
# ANALYTICS DATA
# ============================================================

def analytics_rows(
    db: Session,
):

    engineers = (
        db.query(Engineer)
        .filter(
            Engineer.active == True
        )
        .order_by(
            Engineer.name
        )
        .all()
    )

    result = []

    for engineer in engineers:

        if (
            str(
                engineer.role or ""
            ).upper()
            == "SUPERVISOR"
        ):
            continue

        applications = build_engineer_applications(
            engineer.id,
            db,
        )

        total = len(applications)

        verified = sum(
            1
            for x in applications
            if x["verification_status"]
            == "Verified"
        )

        pending = sum(
            1
            for x in applications
            if x["verification_status"]
            == "Pending"
        )

        issues = sum(
            1
            for x in applications
            if x["verification_status"]
            == "Issue"
        )

        request_not_initiated = sum(
            1
            for x in applications
            if x["ticket_status"]
            == ARM_REQUEST_NOT_INITIATED
        )

        not_required = sum(
            1
            for x in applications
            if x["ticket_status"]
            == ARM_NOT_REQUIRED
        )

        approval_pending = sum(
            1
            for x in applications
            if x["ticket_status"]
            == ARM_APPROVAL_PENDING
        )

        completed = sum(
            1
            for x in applications
            if x["ticket_status"]
            == ARM_COMPLETED
        )

        progress = (
            round(
                (verified / total) * 100,
                1,
            )
            if total
            else 0
        )

        result.append(
            {
                "id": engineer.id,
                "name": engineer.name,
                "alias": engineer.alias,
                "level": engineer.level,
                "email": engineer.email,

                "total": total,
                "verified": verified,
                "pending": pending,
                "issues": issues,
                "progress": progress,

                "request_not_initiated":
                    request_not_initiated,

                "not_required":
                    not_required,

                "approval_pending":
                    approval_pending,

                "completed":
                    completed,

                "applications":
                    applications,
            }
        )

    return result


# ============================================================
# ANALYTICS PAGE
# ============================================================

@router.get(
    "/analytics/{supervisor_id}",
    response_class=HTMLResponse,
)
def analytics_page(
    supervisor_id: int,
    request: Request,
    db: Session = Depends(get_db),
):

    supervisor = get_supervisor_or_403(
        supervisor_id,
        db,
    )

    return templates.TemplateResponse(
        request=request,
        name="analytics.html",
        context={
            "request": request,
            "supervisor": supervisor,
        },
    )


# ============================================================
# ANALYTICS API
# ============================================================

@router.get(
    "/api/analytics/{supervisor_id}/engineers"
)
def analytics_engineers(
    supervisor_id: int,
    db: Session = Depends(get_db),
):

    get_supervisor_or_403(
        supervisor_id,
        db,
    )

    return analytics_rows(db)


# ============================================================
# ANALYTICS CSV EXPORT
# ============================================================

@router.get(
    "/api/analytics/{supervisor_id}/export.csv"
)
def analytics_csv(
    supervisor_id: int,
    db: Session = Depends(get_db),
):

    get_supervisor_or_403(
        supervisor_id,
        db,
    )

    rows = analytics_rows(db)

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow(
        [
            "Engineer",
            "Alias",
            "Level",
            "Email",
            "Application",
            "Access Status",
            "Verification Status",
            "ARM Request Number",
            "ARM Request Status",
            "Remarks",
            "Last Verified",
        ]
    )

    for engineer in rows:

        for application in engineer["applications"]:

            writer.writerow(
                [
                    engineer["name"],
                    engineer["alias"] or "",
                    engineer["level"] or "",
                    engineer["email"] or "",

                    application["application_name"],

                    application["access_status"] or "",

                    application["verification_status"],

                    application["arm_ticket"] or "",

                    application["ticket_status"],

                    application["remarks"] or "",

                    application["last_verified_date"]
                    or "",
                ]
            )

    data = (
        output
        .getvalue()
        .encode("utf-8-sig")
    )

    return StreamingResponse(
        io.BytesIO(data),

        media_type="text/csv; charset=utf-8",

        headers={
            "Content-Disposition":
                "attachment; "
                "filename=access_compliance_analytics.csv"
        },
    )


# ============================================================
# ANALYTICS EXCEL EXPORT
# ============================================================

@router.get(
    "/api/analytics/{supervisor_id}/export.xlsx"
)
def analytics_xlsx(
    supervisor_id: int,
    db: Session = Depends(get_db),
):

    get_supervisor_or_403(
        supervisor_id,
        db,
    )

    rows = analytics_rows(db)

    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
    )

    workbook = Workbook()

    summary = workbook.active

    summary.title = "Engineer Summary"

    summary_headers = [
        "Engineer",
        "Alias",
        "Level",
        "Email",
        "Total",
        "Verified",
        "Pending",
        "Issues",
        "Progress %",
        "Request Not Initiated",
        "Not Required",
        "Approval Pending",
        "Completed",
    ]

    summary.append(summary_headers)

    for cell in summary[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="2563EB",
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for row in rows:

        summary.append(
            [
                row["name"],
                row["alias"] or "",
                row["level"] or "",
                row["email"] or "",
                row["total"],
                row["verified"],
                row["pending"],
                row["issues"],
                row["progress"],
                row["request_not_initiated"],
                row["not_required"],
                row["approval_pending"],
                row["completed"],
            ]
        )

    details = workbook.create_sheet(
        "Application Details"
    )

    detail_headers = [
        "Engineer",
        "Alias",
        "Application",
        "Access Status",
        "Verification Status",
        "ARM Request Number",
        "ARM Request Status",
        "Remarks",
        "Last Verified",
    ]

    details.append(detail_headers)

    for cell in details[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="7C3AED",
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for row in rows:

        for application in row["applications"]:

            details.append(
                [
                    row["name"],
                    row["alias"] or "",
                    application["application_name"],
                    application["access_status"] or "",
                    application["verification_status"],
                    application["arm_ticket"] or "",
                    application["ticket_status"],
                    application["remarks"] or "",
                    application["last_verified_date"] or "",
                ]
            )

    for sheet in workbook.worksheets:

        for column in sheet.columns:

            width = min(
                max(
                    len(
                        str(
                            cell.value or ""
                        )
                    )
                    for cell in column
                )
                + 2,
                40,
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = width

        sheet.freeze_panes = "A2"

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(
        output,

        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        headers={
            "Content-Disposition":
                "attachment; "
                "filename=access_compliance_analytics.xlsx"
        },
    )


# ============================================================
# ALIAS PAGE
#
# KEEP THIS AT THE VERY BOTTOM
# ============================================================

@router.get(
    "/{alias}",
    response_class=HTMLResponse,
)
def engineer_alias_page(
    alias: str,
    request: Request,
    db: Session = Depends(get_db),
):

    engineer = (
        db.query(Engineer)
        .filter(
            Engineer.alias == alias,
            Engineer.active == True,
        )
        .first()
    )

    if engineer is None:

        raise HTTPException(
            status_code=404,
            detail="Engineer not found",
        )

    return page_for(
        engineer,
        request,
        db,
    )